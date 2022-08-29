import smtplib
import string
import time

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

url_base = 'https://www.moex.com/'
HEAD = ('Дата', 'Курс', 'Время', 'Дата', 'Курс', 'Время', 'Результат')
columns = tuple(string.ascii_uppercase)


def to_excel(data):
    """Send data to ecxel document."""
    try:
        data_table = pd.DataFrame(data, columns=HEAD)
        data_table.to_excel('Currency.xlsx',
                            sheet_name='currency',
                            index=False)
    except Exception:
        print('Ошибка записи докуента в excell')


def find_message():
    """Find notification."""
    try:
        clcik = driver.find_element(By.XPATH, '//*[@id="content_disclaimer"]'
                                    '/div/div/div/div[1]/div/a[1]')
        ActionChains(driver).click(clcik).perform()
    except NoSuchElementException:
        print('Notifications was not finded')


def get_dtable():
    """Recieve datatable from xml."""
    driver.find_element(By.XPATH, '/html/body/div[3]/div[3]/div/div/div[1]/'
                        'div[2]/div/div/div/div[2]/form/div[5]/div[3]/table')
    df = pd.read_html(driver.page_source, decimal=",", thousands=" ")[-1]
    df = df.drop(columns='Курс промежуточного клиринга')
    return df.values.tolist()


def merge_dts(dt_usd, dt_jpy):
    """Merge two datatables."""
    for i in range(0, len(dt_usd)):
        result = round(float(dt_usd[i][1]) / float(dt_jpy[i][1]), 2)
        dt_usd[i] += dt_jpy[i]
        dt_usd[i].append(result)
    return dt_usd


def message_len_rows(len):
    """Recieve info messgae about count rows in datatables."""
    rows = ('строка', 'строки', 'строк')
    if len % 10 == 1 and len % 100 != 11:
        line = 0
    elif 2 <= len % 10 <= 4 and (len % 100 < 10 or len % 100 >= 20):
        line = 1
    else:
        line = 2
    return f'В документе {str(len)} {rows[line]}'


def do_finance_format(len, col):
    wb = openpyxl.load_workbook('currency.xlsx')
    ws = wb.active
    for row in range(2, len+2):
        row = f'{col+(str(row))}'
        ws[row].number_format = ('_-* #,##0.00\\ [$₽-19]_-;\\-* #,##0.00\\ '
                                 '[$₽-19]_-;_-* "-"??\\ [$₽-19]_-;_-@_-')
    wb.save("currency.xlsx")


def exist_general_format(len, col):
    wb = openpyxl.load_workbook('currency.xlsx')
    ws = wb.active
    for row in range(2, len+2):
        row = f'{col+(str(row))}'
        if ws[row].number_format != 'General':
            ws[row].number_format = 'General'
    wb.save("currency.xlsx")


def to_alignment(len):
    wb = openpyxl.load_workbook('currency.xlsx')
    ws = wb.active
    for col in columns[:7]:
        for row in range(2, len+2):
            row = f'{col+(str(row))}'
            cell = ws[row]
            cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save("currency.xlsx")


def get_date():
    """Recieve data from source."""
    clcik = driver.find_element(By.XPATH, '//*[@id="d1day"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.PAGE_UP).perform()
    ActionChains(driver).send_keys(Keys.PAGE_UP).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    clcik = driver.find_element(By.XPATH, '//*[@id="d2day"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
    ActionChains(driver).send_keys(Keys.PAGE_DOWN).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    clcik = driver.find_element(By.XPATH, '//*[@id="d2month"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.UP).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    clcik = driver.find_element(By.XPATH, '//*[@id="currency-rate-container"]'
                                '/form/div[4]/div[2]/div/div[5]/input')
    ActionChains(driver).click(clcik).perform()


def send_email(msg):
    """Send mail to MikAlBelov@Greenatom.ru"""
    try:
        server = smtplib.SMTP_SSL('smtp.gmail.com', 587)
        server.ehlo()
        server.login('test@gmail.com', 'test')
        server.sendmail('test@gmail.com', 'MikAlBelov@Greenatom.ru',
                        msg.as_string())
        server.quit()
    except Exception as e:
        print(f'Ошибка при отправке почты. Сообщение: {msg}. Error: {e}')


if __name__ == "__main__":
    chrome_options = webdriver.ChromeOptions()
    prefs = {'profile.default_content_setting_values.notifications': 2}
    chrome_options.add_experimental_option('prefs', prefs)
    driver = webdriver.Chrome(chrome_options=chrome_options)
    driver.get(url_base)
    clcik = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div/div/'
                                'div[1]/div/div[2]/button')
    ActionChains(driver).click(clcik).perform()
    clcik = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/div/div/'
                                'div[1]/div/div[2]/div[4]/div[2]/div[1]/a')
    ActionChains(driver).click(clcik).perform()
    find_message()
    clcik = driver.find_element(By.XPATH,
                                '/html/body/div[3]/div[2]/div/div/div[1]/div/'
                                'div[2]/div[4]/div[2]/div[1]/div/div[3]/a')
    ActionChains(driver).click(clcik).perform()
    time.sleep(2)
    find_message()
    driver.set_window_size(1920, 1080)
    clcik = driver.find_element(By.XPATH, '//*[@id="ctl00_frmLeftMenuWrap"]'
                                '/div/div/div/div[2]/div[13]/a')
    ActionChains(driver).click(clcik).perform()
    get_date()
    df_list_usd = get_dtable()
    clcik = driver.find_element(By.XPATH, '//*[@id="ctl00_PageContent_'
                                'CurrencySelect"]')
    ActionChains(driver).click(clcik).perform()
    ActionChains(driver).send_keys(Keys.DOWN).perform()
    ActionChains(driver).send_keys(Keys.DOWN).perform()
    ActionChains(driver).send_keys(Keys.DOWN).perform()
    ActionChains(driver).send_keys(Keys.DOWN).perform()
    ActionChains(driver).send_keys(Keys.DOWN).perform()
    ActionChains(driver).send_keys(Keys.ENTER).perform()
    get_date()
    df_list_jpy = get_dtable()
    all_data = merge_dts(df_list_usd, df_list_jpy)
    to_excel(all_data)
    len_rows = len(all_data)
    do_finance_format(len_rows, columns[1])
    do_finance_format(len_rows, columns[4])
    exist_general_format(len(all_data), columns[6])
    to_alignment(len_rows)
    msg = message_len_rows((len_rows + 1))
    send_email(msg)
    driver.quit()
